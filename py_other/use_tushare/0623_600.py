from pathlib import Path
import tushare as ts
import time
import numpy as np
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
import pandas as pd

# 四个行业
# industries = ['电子', '通信', '传媒', '计算机']
industries = ['电子2']

# 保存四个行业公司股票代码的excel路径
excel_path = './四行业数据.xlsx'


class Run:
    def __init__(self, token, start, end):
        self.token = token
        self.start = start
        self.end = end

    def run(self):
        infos = self.preprocess()
        pro = ts.pro_api(token=self.token)
        wb = Workbook()

        # 存放原始数据的工作簿
        raw_wb = Workbook()
        # 遍历每个行业
        for idx, (industry_name, codes) in enumerate(infos.items()):
            print(f'正在处理 {industry_name} ...')
            ws = wb.create_sheet(industry_name, idx + 1)
            raw_ws = raw_wb.create_sheet(industry_name, idx + 1)
            industry_closes, trade_dates, dfs = self.get_datas(start='2010-05-01',
                                                               end='2023-04-30',
                                                               pro=pro,
                                                               codes=codes)
            # 保存爬到的原始数据
            self.write_raw_data(dfs, raw_wb, raw_ws)
            # [公司数, 日股票信息], 如 [100, 3140] 就表示有100家公司，每家公司的股票信息是3140个
            # print(np.array(industry_closes).shape)

            industry_closes2, _, _ = self.get_datas(start='2009-03-01',
                                                    end='2010-04-30',
                                                    pro=pro,
                                                    codes=codes)

            num1 = len(industry_closes[0])
            # 合并数据
            for idx, c in enumerate(industry_closes2):
                for c_ in c:
                    industry_closes[idx].append(c_)

            # 计算行业的净新高占比
            npnmns, npnmns_infos = self.cal_nh_nl(industry_closes, num1)
            print(len(npnmns), len(npnmns_infos))
            self.pic(npnmns, industry_closes, trade_dates, industry_name)
            tanlan, konghuang = self.cal_acc(npnmns_infos, trade_dates)

            self.proce_excel(tanlan=tanlan,
                             konghuang=konghuang,
                             wb=wb,
                             ws=ws,
                             industry_name=industry_name)

            # break

    @staticmethod
    def write_raw_data(infos, wb, ws):
        """
        保存爬到的所有原始数据
        Returns:

        """
        ws.append(['ts_code', 'trade_date', 'open', 'high'
                   'low', 'close', 'pre_close', 'change',
                   'pct_chg', 'vol', 'amount'])
        for info in infos:
            ts_code = info['ts_code']
            trade_date = info['trade_date']
            open_ = info['open']
            high = info['high']
            low = info['low']
            close = info['close']
            pre_close = info['pre_close']
            change = info['change']
            pct_chg = info['pct_chg']
            vol = info['vol']
            amount = info['amount']

    # 第一步，定义计算净新高占比的函数
    @staticmethod
    def cal_nh_nl(industry_closes, num):
        """
        计算行业净新高占比
        industry_closes : [公司数量, 股票信息数]
        创年度新高 : 1
        创年度新低 : 2
        否则 : 0
        Returns:

        """
        counts = []
        max_len = len(industry_closes[0])
        for idx, c in enumerate(industry_closes):
            temp = []
            for i in range(num):
                # 创年度新高
                if c[i] > max(c[i + 5: min(i + 52 * 5, max_len)]):
                    temp.append(1)

                # 创年度新低
                elif c[i] < min(c[i + 5: min(i + 52 * 5, max_len)]):
                    temp.append(2)

                else:
                    temp.append(0)
            counts.append(temp)

        # 统计上市公司的数量
        shangshi_nums = []
        for i in np.array(industry_closes).T:
            temp_num = np.where(i > 0)[0].size
            shangshi_nums.append(temp_num)

        npnmns = []
        npnmns_infos = []
        counts = np.array(counts).T
        # 贪婪:1 乐观:2 正常区间:3 悲观:4 恐惧:5
        for idx, c in enumerate(counts):
            nh = np.where(c == 1)[0].size  # 创年度新高的个股数
            nl = np.where(c == 2)[0].size  # 创年度新低的个股数
            if shangshi_nums[idx] == 0:
                npnmns.append(0)
                npnmns_infos.append(0)
            else:
                npnmn = (nh - nl) / shangshi_nums[idx]
                npnmns.append(npnmn)
                if nh + nl >= 40:
                    if npnmn >= 0.3:
                        npnmns_infos.append(1)
                    elif 0.2 <= npnmn < 0.3:
                        npnmns_infos.append(2)
                    elif -0.2 < npnmn < 0.2:
                        npnmns_infos.append(3)
                    elif -0.3 < npnmn < -0.2:
                        npnmns_infos.append(4)
                    elif npnmn <= -0.3:
                        npnmns_infos.append(5)
                    else:
                        npnmns_infos.append(0)
                else:
                    if npnmn >= 0.4:
                        npnmns_infos.append(1)
                    elif 0.3 <= npnmn < 0.4:
                        npnmns_infos.append(2)
                    elif -0.3 < npnmn < 0.3:
                        npnmns_infos.append(3)
                    elif -0.4 < npnmn < -0.3:
                        npnmns_infos.append(4)
                    elif npnmn <= -0.4:
                        npnmns_infos.append(5)
                    else:
                        npnmns_infos.append(0)

        return npnmns, npnmns_infos

    def get_datas(self, start, end, pro, codes):
        industry_closes = []
        trade_dates = []
        dfs = []
        # 遍历每个公司的股票代码
        for code in codes:
            # df = ts.get_hist_data(code=code, start=start, end=end)
            df = pro.daily(ts_code=code, start_date=start, end_date=end)

            df = df.to_dict()
            print(df.keys())

            dfs.append(df)

            # 收盘价
            closes = df['close']

            # 交易日
            trade_date = df['trade_date']

            industry_closes.append(closes)
            trade_dates.append(trade_date)

            # break
        assert len(industry_closes) == len(trade_dates), 'should be same'
        industry_closes, trade_dates = self.alignment(industry_closes, trade_dates)

        new_industry_closes, new_industry_highs, new_industry_lows = [], [], []
        for c in industry_closes:
            temp = []
            for k in c.keys():
                temp.append(c[k])

            new_industry_closes.append(temp)

        # self.cal_nh_nl(new_industry_closes, new_industry_highs, new_industry_lows)
        return new_industry_closes, trade_dates, dfs

    @staticmethod
    def alignment(industry_closes, trade_dates):
        """
        同一行业下的公司，因为上市时间有早晚，可能会导致同一时间段内的股票信息数量不一致
        所以该函数就是做数据对齐操作
        Returns:

        """
        new_industry_closes = []
        new_trade_dates = []
        arr = [len(c) for c in industry_closes]
        max_value = int(np.max(arr))
        max_value_idx = int(np.argmax(arr))
        for c, d in zip(industry_closes, trade_dates):
            if len(c) == max_value:
                new_industry_closes.append(c)
                new_trade_dates.append(d)
            # 补充缺失信息
            else:
                temp = {}
                temp2 = {}
                for k in industry_closes[max_value_idx].keys():
                    if k in c.keys():
                        temp[k] = c[k]
                        temp2[k] = d[k]
                    else:
                        temp[k] = 0
                        temp2[k] = trade_dates[max_value_idx][k]
                new_industry_closes.append(temp)
                new_trade_dates.append(temp2)

        return new_industry_closes, new_trade_dates

    @staticmethod
    def preprocess():
        """
        预处理excel文件
        Returns:
            dict {行业名称: [股票代码, ...]}
        """
        infos = {}
        wb = load_workbook(filename=excel_path)
        for i in industries:
            infos[i] = []
            ws = wb[i]
            for row_idx in range(2, ws.max_row + 1):
                loc = 'E' + str(row_idx)
                v = str(ws[loc].value).rjust(8, '0')
                if i == '计算机' or i == '计算机2':
                    v = v[2:] + '.' + v[:2]
                infos[i].append(v)
        # for k, v in infos.items():
        #     print(k, len(v))
        wb.close()
        return infos

    @staticmethod
    def pic(npnmns, industry_closes, trade_dates, industry_name):
        """
        绘制图表并保存
        trade_dates: 交易日
        Returns:

        """
        closes = np.array(industry_closes).T
        closes = np.sum(closes, axis=-1).tolist()[:len(npnmns)]

        # trade_dates = [v for _, v in trade_dates[0].items()]
        # 把横坐标的日期改成索引
        trade_dates = range(len(trade_dates[0]))

        trade_dates = trade_dates[::-1]

        fig, ax1 = plt.subplots()
        plt.xticks(np.arange(0, len(trade_dates), step=365 * 2))

        l1 = ax1.plot(trade_dates, closes, color="blue", linewidth=1.0)
        ax1.set_xlabel(" ")
        ax1.set_ylabel("¥")

        l2 = ax2 = ax1.twinx()
        ax2.plot(trade_dates, npnmns, color="brown", linewidth=1.0)
        ax2.set_ylabel("%")

        fig.legend(
            labels=["industry", "(NH-NL)%"],
            loc="lower right",
            fontsize=6)
        save_path = industry_name + '.png'
        plt.savefig(save_path)

    @staticmethod
    def cal_acc(npnmns_infos, trade_dates):
        """
        计算净新高占比提示某行业的准确率
        Returns:
        """
        years = ['2010', '2011', '2012', '2013', '2014',
                 '2015', '2016', '2017', '2018', '2019',
                 '2020', '2021', '2022', '2023']
        months = ['01', '02', '03', '04', '05', '06',
                  '07', '08', '09', '10', '11', '12']
        delta = []
        for y in years:
            for m in months:
                de = y + m
                delta.append(de)
        trade_dates = [v for _, v in trade_dates[0].items()]

        info1 = {}
        info2 = {}
        # 贪婪:1 乐观:2 正常区间:3 悲观:4 恐惧:5
        for t, npnmn in zip(trade_dates[::-1], npnmns_infos[::-1]):
            if npnmn == 1:
                if t[:6] not in info1.keys():
                    info1[t[:6]] = []
                info1[t[:6]].append(npnmn)
            elif npnmn == 5:
                if t[:6] not in info2.keys():
                    info2[t[:6]] = []
                info2[t[:6]].append(npnmn)
            else:
                continue

        new_info1 = {}
        new_info2 = {}
        for k, v in info1.items():
            new_info1[k] = len(v)

        for k, v in info2.items():
            new_info2[k] = len(v)

        print(new_info1.keys(), len(new_info1), new_info2.keys(), len(new_info2))
        return new_info1, new_info2

    @staticmethod
    def proce_excel(tanlan, konghuang, wb, ws, industry_name):
        ws['A1'] = '(NH-NL)%提示贪婪'
        ws['B1'] = '维持天数'
        for idx, (k, v) in enumerate(tanlan.items()):
            loc1 = 'A' + str(idx + 2)
            loc2 = 'B' + str(idx + 2)
            ws[loc1] = k
            ws[loc2] = str(v)

        ws['D1'] = '(NH-NL)%提示恐慌'
        ws['E1'] = '维持天数'
        for idx, (k, v) in enumerate(konghuang.items()):
            loc1 = 'D' + str(idx + 2)
            loc2 = 'E' + str(idx + 2)
            ws[loc1] = k
            ws[loc2] = str(v)

        img_name = industry_name + '.png'
        img = Image(img_name)
        ws.add_image(img, 'G3')

        wb.save('info.xlsx')


def main():
    start = time.time()
    t = '818670fa68bc204c217143cdb75efeae1986031841ff8ca2c6a855bd'
    s = '20100501'
    e = '20230430'
    ru = Run(token=t, start=s, end=e)
    ru.run()
    cost_time = time.time() - start
    print(f"耗时 {cost_time:.2f} s")


def cal_sy_hc():
    """
    计算收益率和回撤率
    Returns:

    """
    # 准备行业指数的历史价格数据和净新高占比指标数据 gb18030
    price_data = pd.read_csv('price_data.csv', encoding='gb18030')
    indicator_data = pd.read_csv('indicator_data.csv', encoding='gb18030')
    # 合并价格数据和指标数据
    data = pd.merge(price_data, indicator_data, on='日期')

    # 设置策略参数
    stop_loss = 0.01  # 止损比例
    holding_period = 30  # 持有期

    # 初始化策略指标
    cumulative_returns = 1.0
    max_drawdown = 1.0

    # 遍历每个交易日进行回测
    for i in range(len(data)):
        # 判断贪婪信号
        if data['净新高占比'].iloc[i] > 0:
            # 判断是否回到阈值内
            if data['收盘价'].iloc[i] <= data['阈值'].iloc[i]:
                # 做空交易
                entry_price = data['开盘价'].iloc[i + 1]
                stop_loss_level = data['最高价'].iloc[i - 5:i].max()  # 前一周区间最高点
                exit_price = data['收盘价'].iloc[i + holding_period]

                # 判断是否触发止损
                if exit_price > stop_loss_level:
                    exit_price = entry_price  # 将止损位移至成本价

                cumulative_returns *= exit_price / entry_price
                max_drawdown = min(max_drawdown, cumulative_returns)

        # 判断恐慌信号
        elif data['净新高占比'].iloc[i] < 0:
            # 判断是否回到阈值内
            if data['收盘价'].iloc[i] >= data['阈值'].iloc[i]:
                # 做多交易
                entry_price = data['开盘价'].iloc[i + 1]
                stop_loss_level = data['最低价'].iloc[i - 5:i].min()  # 前一周区间最低点
                exit_price = data['收盘价'].iloc[i + holding_period]

                # 判断是否触发止损
                if exit_price < stop_loss_level:
                    exit_price = entry_price  # 将止损位移至成本价

                cumulative_returns *= exit_price / entry_price
                max_drawdown = min(max_drawdown, cumulative_returns)

    # 计算策略表现
    cumulative_returns = (cumulative_returns - 1) * 100  # 转化为百分比收益
    max_drawdown = (1 - max_drawdown) * 100  # 转化为百分比最大回撤

    print('累计收益率：', cumulative_returns, '%')
    print('最大回撤：', max_drawdown, '%')


# 绘图
def pic(npnmns, industry_closes, trade_dates, industry_name):
    """
    绘制图表并保存
    trade_dates: 交易日
    Returns:

    """
    closes = np.array(industry_closes).T
    closes = np.sum(closes, axis=-1).tolist()[:len(npnmns)]

    # trade_dates = [v for _, v in trade_dates[0].items()]
    # 把横坐标的日期改成索引
    trade_dates = range(len(trade_dates[0]))

    trade_dates = trade_dates[::-1]

    fig, ax1 = plt.subplots()
    plt.xticks(np.arange(0, len(trade_dates), step=365 * 2))

    l1 = ax1.plot(trade_dates, closes, color="blue", linewidth=1.0)
    ax1.set_xlabel(" ")
    ax1.set_ylabel("¥")

    l2 = ax2 = ax1.twinx()
    ax2.plot(trade_dates, npnmns, color="brown", linewidth=1.0)
    ax2.set_ylabel("%")

    fig.legend(
        labels=["industry", "(NH-NL)%"],
        loc="lower right",
        fontsize=6)
    save_path = industry_name + '.png'
    plt.savefig(save_path)


def test02():
    pro = ts.pro_api(token='818670fa68bc204c217143cdb75efeae1986031841ff8ca2c6a855bd')
    df = pro.daily(ts_code='000001.SZ', start='20200720', end='20200723')
    datas = pro.stock_basic()
    print(df)
    print(datas)


if __name__ == "__main__":
    # main()
    # cal_sy_hc()
    test02()
