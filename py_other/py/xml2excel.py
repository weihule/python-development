#
#  XML文件转换成Excel文件
#
#  版本：1.0
#
#  作者：Witton Bell
#  E_Mail:witton@163.com
#
#
#  功能描述：
#
#  自动检测源文件编码是否为UTF8，如果不是，则使用GB18030进行解码，并转换成UTF8编码的文件，再进行XML源文件的读取
#
#  自动提取XML中的标签作为Excel的标题列，支持嵌套与重复，但是所有列不能超过EXCEL支持的最大列（16384列)
#
#  仅支持XML的第一层只有一个标签，
#  第二层标签可以只有一个标签（允许重复），此时第一层标签作为Excel的WorkSheet，第二层标签及其嵌套作为一行数据）
#  第二层标签也可以有多个标签（允许重复），此时第二层标签作为Excel的WorkSheet，第三层标签及其嵌套作为一行数据）
#
#  由于openpyxl默认为Excel的xlsx格式，所以本脚本的所有Excel文件格式都为xlsx
#
#
import openpyxl
from openpyxl.styles import Alignment
import xml.dom.minidom
from xml.dom.minidom import Document
from openpyxl.styles import Border, Side
import chardet
import os


class XMLNode:
    def __init__(self):
        self.name = ""
        self.properties = []
        self.child = []
        self.layer = 0
        self.index = 0
        self.parent = None
        self.node_info = None


class XMLNodeInfo:
    def __init__(self, node):
        self.map_properties = {}
        self.map_child = {}
        self.max_index = 0
        self.layer = node.layer
        self.name = node.name

        self.register(node)

    def register(self, node):
        for k in node.properties:
            if self.map_properties.get(k[0]) is None:
                self.map_properties[k[0]] = self.map_properties.__len__()

        for ch in node.child:
            v = self.map_child.get(ch.name)
            if v is None:
                self.map_child[ch.name] = ch.node_info

        if node.index > self.max_index:
            self.max_index = node.index


class XMLReader:
    # is_auto_convert2utf8：是否自动转换为UTF8编码
    # is_convert_to_original_file：如果转换为UTF8，是否覆盖原文件
    def __init__(self, file_path, is_auto_convert2utf8, is_convert_to_original_file):
        self.__root = []
        self.__map_node_info = {}
        if is_auto_convert2utf8:
            is_tmp_file, tmp_file_path = self.__convert_to_utf8(file_path)
            fd = xml.dom.minidom.parse(tmp_file_path)
            if is_tmp_file:
                if is_convert_to_original_file:
                    os.remove(file_path)
                    os.rename(tmp_file_path, file_path)
                else:
                    os.remove(tmp_file_path)
        else:
            fd = xml.dom.minidom.parse(file_path)
        index = 0
        for child in fd.childNodes:
            if child.nodeType != xml.dom.Node.ELEMENT_NODE:
                continue

            self.__read_node(child, self.__root, index, None)
            index += 1

    def get_root_node(self):
        return self.__root

    @staticmethod
    def __convert_to_utf8(file_path):
        fd = open(file_path, "rb")
        fd.seek(0, 2)
        size = fd.tell()
        if size > 1024 * 1024:
            size = 1024 * 1024
        fd.seek(0, 0)
        text = fd.read(size)
        ret = chardet.detect(text)
        if ret['encoding'].lower().find("utf-8") != -1:
            return False, file_path

        tmp_file = file_path + ".tmp"
        file = open(tmp_file, "w", encoding="utf-8", newline="\n")
        fd.seek(0, 0)
        line = fd.readline()
        while line.__len__() > 0:
            file.write(line.decode("gb18030"))
            line = fd.readline()
        file.close()
        fd.close()

        return True, tmp_file

    @staticmethod
    def __get_attrib(node, rc):
        if node._attrs is None:
            return
        for key in node._attrs:
            v = node._attrs[key]._value
            rc.append([key, v])

    def __read_node(self, node, root, index, parent, layer=1):
        xml_node = XMLNode()
        xml_node.name = node.nodeName
        xml_node.layer = layer
        xml_node.index = index
        xml_node.parent = parent
        self.__get_attrib(node, xml_node.properties)
        i = 0
        for child in node.childNodes:
            if child.nodeType != xml.dom.Node.ELEMENT_NODE:
                continue

            self.__read_node(child, xml_node.child, i, xml_node, layer + 1)
            i += 1
        root.append(xml_node)
        self.__register(xml_node)

    def __register(self, node):
        key = node.name + str(node.layer)
        nd = self.__map_node_info.get(key)
        if nd is None:
            nd = XMLNodeInfo(node)
            node.node_info = nd
            self.__map_node_info[key] = nd
        else:
            nd.register(node)
            node.node_info = nd


class XMLWriter:
    def __init__(self, xml_file_path, xml_node_list):
        doc = Document()
        for node in xml_node_list:
            ele = self.__write_node(node, doc)
            doc.appendChild(ele)

        f = open(xml_file_path, 'w', encoding='utf-8')
        doc.writexml(f, indent='\t', addindent='\t', encoding='utf-8', newl='\n')

    def __write_node(self, node, doc):
        ele = doc.createElement(node.name)
        for prop in node.properties:
            ele.setAttribute(prop[0], prop[1])

        for child in node.childs:
            ret = self.__write_node(child, doc)
            ele.appendChild(ret)

        return ele


class XmlToXls:
    # read_from_xml_file_path：XML源文件完整路径
    # save_to_xls_file_path：保存转换后的Excel文件完整路径
    # is_auto_convert2utf8：是否自动转换为UTF8编码
    # is_convert_to_original_file：如果转换为UTF8，是否覆盖原文件
    # is_merge_head：如果表头有多行，是否合并层属关系的表头
    # is_border_cell：是否添加单元格的边框
    # is_alignment_center：单元格是否居中对齐
    def __init__(self, read_from_xml_file_path, save_to_xls_file_path,
                 is_auto_convert2utf8=True, is_convert_to_original_file=False,
                 is_merge_head=False, is_border_cell=False, is_alignment_center=True):
        try:
            self.__is_merge_head = is_merge_head
            if is_alignment_center:
                self.__alignment = Alignment(horizontal='center', vertical='center')
            else:
                self.__alignment = None
            if is_border_cell:
                side = Side(border_style='thin', color='000000')
                self.__border = Border(side, side, side, side, side)
            else:
                self.__border = None
            wb = openpyxl.Workbook()
            wb.encoding = 'utf-8'
            for sh in wb.worksheets:
                wb.remove_sheet(sh)

            reader = XMLReader(read_from_xml_file_path,
                               is_auto_convert2utf8,
                               is_convert_to_original_file)
            self.__write(reader.get_root_node(), wb)
            wb.save(save_to_xls_file_path)
            wb.close()
        except Exception as e:
            print(e)

    def __write(self, xml_node, wb):
        self.__map_field = {}
        for node in xml_node:
            if node.node_info.map_child.__len__() == 1:
                self.__start_layer = 1
                self.__write_sheet(node, wb)
            else:
                self.__start_layer = 2
                for child in node.child:
                    self.__write_sheet(child, wb)

    def __write_sheet(self, node, wb):
        sh = wb.create_sheet(node.name)
        self.__write_head(node.node_info, sh)
        self.__write_data(node, sh)

    def __write_head(self, node, sh):
        for key in node.map_child:
            col = 1
            child = node.map_child[key]
            for k in child.map_properties:
                c = child.map_properties[k]
                self.__write_cell(sh, child.layer - self.__start_layer, c + col, k)
            col += child.map_properties.__len__()
            self.__write_head_ext(child, sh, col)
            break

    def __write_head_ext(self, node, sh, col):
        for key in node.map_child:
            child = node.map_child[key]
            num = child.map_properties.__len__()
            for i in range(0, child.max_index + 1):
                if col > 16384:
                    raise Exception("超过EXCEL最大列数(16384列)限制，转换失败")

                old_col = col
                self.__write_cell(sh, child.layer - self.__start_layer - 1, col, child.name)
                for k in child.map_properties:
                    c = child.map_properties[k]
                    self.__write_cell(sh, child.layer - self.__start_layer, c + col, k)
                col += num
                col = self.__write_head_ext(child, sh, col)
                if self.__is_merge_head:
                    merge_row = child.layer - self.__start_layer - 1
                    sh.merge_cells(start_row=merge_row, end_row=merge_row,
                                   start_column=old_col, end_column=col - 1)
        return col

    def __write_data(self, node, sh):
        row = sh.max_row + 1
        sh.freeze_panes = sh.cell(row, 1)
        for child in node.child:
            col = 1
            self.__write_data_ext(child, sh, row, col)
            row += 1

    def __write_data_ext(self, node, sh, row, col):
        m = node.node_info.map_properties
        max_prop_num = m.__len__()
        for prop in node.properties:
            c = m[prop[0]]
            self.__write_cell(sh, row, c + col, prop[1])
        col += max_prop_num

        map_child = {}
        for child in node.child:
            if map_child.get(child.name) is None:
                map_child[child.name] = 1
            else:
                map_child[child.name] += 1
            col = self.__write_data_ext(child, sh, row, col)

        for key in node.node_info.map_child:
            child = node.node_info.map_child[key]
            all_count = child.max_index + 1
            count = map_child.get(key)
            if count is not None:
                all_count -= count

            for i in range(0, all_count):
                col += child.map_properties.__len__()

        return col

    def __write_cell(self, sh, row, col, value):
        if value.isdigit():
            cell = sh.cell(row, col, int(value))
        else:
            cell = sh.cell(row, col, value)
        cell.alignment = self.__alignment
        cell.border = self.__border

