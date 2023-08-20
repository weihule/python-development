from openpyxl import Workbook, load_workbook
from openpyxl.utils import FORMULAE
import datetime
import sys
from pathlib import Path
import cv2


def main():
    wb = Workbook()
    ws = wb.active

    ws.append([22, 63])
    ws.append([11, 88])
    ws.append([15, 68])

    ws["c2"] = "=SUM(A2, B2)"   # 求和
    ws["d2"] = "=AVERAGE(A2: B2)"  # 求平均值

    ws.column_dimensions['A'].width = 20.0  # 调整列A宽
    ws.row_dimensions[1].height = 40  # 调整行1高

    wb.save('test.xlsx')

    ws1 = wb.create_sheet("my_sheet1", 1)


def main2():
    myExcel = load_workbook('省会表.xlsx')  # 获取表格文件
    # 获取指定的表单
    mySheet = myExcel.get_sheet_by_name('Sheet1')


def test():
    p = Path(r"D:\desktop\test.txt")
    p.unlink(missing_ok=True)


if __name__ == "__main__":
    # main()
    test()


