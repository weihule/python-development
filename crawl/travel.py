from pathlib import Path
import requests
from lxml import etree
from openpyxl import Workbook, load_workbook
import json

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.63 Safari/537.36 Edg/102.0.1245.30'
}
save_root = Path(r'D:\Desktop\provinces')


def main():
    url = "https://www.maigoo.com/goomai/197792.html"
    response = requests.get(url=url, headers=headers)
    page_text = response.text

    # xpath解析网页源码
    page_text = etree.HTML(page_text)

    provinces = page_text.xpath(
        '//div[@class="md_citiaothemetop10 md_citiaothemetop10_clos"]/div')[0].xpath('.//li')

    provinces = page_text.xpath(
        '//div[@class="md_citiaothemetop10 md_citiaothemetop10_clos"]/div')

    print(provinces)

    infos = {}

    for province in provinces:
        # 省份名称
        prov_name = province.xpath('.//span/text()')[0].split('A级')[0].replace('国家', '')

        # 初始化
        infos[prov_name] = []

        if prov_name in ['贵州', '北京', '黑龙江', '江西', '山西']:
            continue

        # 省份对应的url
        prov_url = province.xpath('.//a[@target="_blank"]/@href')[0]

        # 请求各省对应的url，获取对应的网页源码
        prov_response = requests.get(url=prov_url, headers=headers)
        prov_page_text = prov_response.text
        prov_page_text = etree.HTML(prov_page_text)

        attractions = prov_page_text.xpath('.//div[@class="md_6021 modelbox tleft"]')
        for attraction_infos in attractions:
            # 遍历每个城市中的景区
            for idx, attraction in enumerate(attraction_infos.xpath('.//tr[@class="li font14"]')):
                # 1. 景区名称
                if len(attraction) != 4:
                    break
                attraction_name = attraction[1].xpath('./text()')
                if not attraction_name:
                    attraction_name = attraction[1].xpath('.//a[@target="_blank"]/text()')[0]

                if isinstance(attraction_name, list) and len(attraction_name) > 0:
                    attraction_name = attraction_name[0]
                # 2. 景区所在地
                location = attraction[2].xpath('./text()')[0]

                # 3. 景区级别和评级时间
                attraction_level = attraction[3].xpath('./text()')[0]. \
                    replace('（', '(').replace('）', ')').replace(' ', '')

                infos[prov_name].append([attraction_name, location, attraction_level])

    # ---------- 山西 ----------
    shanxi = provinces[16]
    prov_name = '山西'
    prov_url = shanxi.xpath('.//a[@target="_blank"]/@href')[0]

    # 请求山西省对应的url，获取对应的网页源码
    prov_response = requests.get(url=prov_url, headers=headers)
    prov_page_text = prov_response.text
    prov_page_text = etree.HTML(prov_page_text)

    city_names = prov_page_text.xpath('.//div[@class="pagecolor28"]/div[@class="md_2053 modelbox tcolor34 tleft "]')[:-1]
    city_infos = prov_page_text.xpath('.//div[@class="pagecolor28"]/div[@class="md_6021 modelbox tleft"]')

    # 遍历山西省下属每个城市及城市中的景区
    for city_name, city_info in zip(city_names, city_infos):
        city_name = city_name.xpath('.//div[@class="mod_txt fcolor"]/text()')[0].replace(' ', '').split('A级')[0]

        # 遍历每个城市中的景区
        for idx, attraction in enumerate(city_info.xpath('.//tr[@class="li font14"]')):
            # 1. 景区名称
            attraction_name = attraction[1].xpath('./text()')
            if not attraction_name:
                attraction_name = attraction[1].xpath('.//a[@target="_blank"]/text()')[0]

            if isinstance(attraction_name, list) and len(attraction_name) > 0:
                attraction_name = attraction_name[0]
            # 2. 景区所在地
            location = city_name

            # 3. 景区级别和评级时间
            attraction_level = attraction[2].xpath('./text()')[0].\
                replace('（', '(').replace('）', ')').replace(' ', '')

            infos[prov_name].append([attraction_name, location, attraction_level])
    # ---------- 山西 ----------

    # ---------- 江西 ----------
    shanxi = provinces[9]
    prov_name = '江西'
    prov_url = shanxi.xpath('.//a[@target="_blank"]/@href')[0]

    # 请求江西省对应的url，获取对应的网页源码
    prov_response = requests.get(url=prov_url, headers=headers)
    prov_page_text = prov_response.text
    prov_page_text = etree.HTML(prov_page_text)

    city_names = prov_page_text.xpath('.//div[@class="pagecolor15"]/div[@class="md_2027 modelbox tcolor34 tcenter"]')[:-1]
    city_infos = prov_page_text.xpath('.//div[@class="pagecolor15"]/div[@class="md_6021 modelbox tleft"]')

    # 遍历江西省下属每个城市及城市中的景区
    for city_name, city_info in zip(city_names, city_infos):
        city_name = city_name.xpath('.//div[@class="mod_word fcolor font28 line18em"]/text()')[0].replace(' ', '').split('A级')[0]

        # 遍历每个城市中的景区
        for idx, attraction in enumerate(city_info.xpath('.//tr[@class="li font14"]')):
            # 1. 景区名称
            attraction_name = attraction[1].xpath('./text()')
            if not attraction_name:
                attraction_name = attraction[1].xpath('.//a[@target="_blank"]/text()')[0]

            if isinstance(attraction_name, list) and len(attraction_name) > 0:
                attraction_name = attraction_name[0]
            # 2. 景区所在地
            location = city_name

            # 3. 景区级别和评级时间
            attraction_level = attraction[2].xpath('./text()')[0].\
                replace('（', '(').replace('）', ')').replace(' ', '')

            infos[prov_name].append([attraction_name, location, attraction_level])
    # ---------- 江西 ----------

    # ---------- 黑龙江 ----------
    heilongjiang = provinces[13]
    prov_name = '黑龙江'
    prov_url = heilongjiang.xpath('.//a[@target="_blank"]/@href')[0]

    # 请求黑龙江省对应的url，获取对应的网页源码
    prov_response = requests.get(url=prov_url, headers=headers)
    prov_page_text = prov_response.text
    prov_page_text = etree.HTML(prov_page_text)

    city_names = prov_page_text.xpath('.//div[@class="pagecolor1"]/div[@class="md_2027 modelbox tcenter"]')[:-1]
    city_infos = prov_page_text.xpath('.//div[@class="pagecolor1"]/div[@class="md_6021 modelbox tleft"]')

    # 遍历黑龙江省下属每个城市及城市中的景区
    for city_name, city_info in zip(city_names, city_infos):
        city_name = city_name.xpath('.//div[@class="mod_word fcolor font28 line18em"]/text()')[0].replace(' ', '').split('A级')[0]

        # 遍历每个城市中的景区
        for idx, attraction in enumerate(city_info.xpath('.//tr[@class="li font14"]')):
            # 1. 景区名称
            attraction_name = attraction[1].xpath('./text()')
            if not attraction_name:
                attraction_name = attraction[1].xpath('.//a[@target="_blank"]/text()')[0]

            if isinstance(attraction_name, list) and len(attraction_name) > 0:
                attraction_name = attraction_name[0]
            # 2. 景区所在地
            location = city_name

            # 3. 景区级别和评级时间
            attraction_level = attraction[2].xpath('./text()')[0].\
                replace('（', '(').replace('）', ')').replace(' ', '')

            infos[prov_name].append([attraction_name, location, attraction_level])
    # ---------- 黑龙江 ----------

    # ---------- 天津 ----------
    tianjin = provinces[21]
    prov_name = '天津'
    prov_url = tianjin.xpath('.//a[@target="_blank"]/@href')[0]

    # 请求山西省对应的url，获取对应的网页源码
    prov_response = requests.get(url=prov_url, headers=headers)
    prov_page_text = prov_response.text
    prov_page_text = etree.HTML(prov_page_text)

    attractions = prov_page_text.xpath('.//tr[@class="li font14"]')
    # 遍历天津市的景区
    for idx, attraction in enumerate(attractions):
        # 1. 景区名称
        attraction_name = attraction[1].xpath('./text()')
        if not attraction_name:
            attraction_name = attraction[1].xpath('.//a[@target="_blank"]/text()')[0]

        if isinstance(attraction_name, list) and len(attraction_name) > 0:
            attraction_name = attraction_name[0]
        # 2. 景区所在地
        location = attraction[2].xpath('./text()')[0]

        # 3. 景区级别和评级时间
        attraction_level = attraction[3].xpath('./text()')[0]. \
            replace('（', '(').replace('）', ')').replace(' ', '')

        infos[prov_name].append([attraction_name, location, attraction_level])
    # ---------- 天津 ----------

    # ---------- 贵州 ----------
    guizhou_url_4a = "https://www.guizhou.gov.cn/dcgz/hkgz/rmjd/202109/t20210914_70370942.html"
    guizhou_url_5a = "https://www.guizhou.gov.cn/dcgz/hkgz/rmjd/202109/t20210914_70370943.html"
    urls = [guizhou_url_4a, guizhou_url_5a]
    for guizhou_idx, i in enumerate(urls):
        response = requests.get(url=i, headers=headers)
        # 写成这种格式，防止中文乱码
        page_text = etree.HTML(response.content, parser=etree.HTMLParser(encoding='utf8'))
        # 去掉第一个
        attractions = page_text.xpath('.//div[@class="ue_table"]//tr')[1:]
        for attraction in attractions:
            name = attraction[2].xpath('./text()')[0]       # 景区名称
            location = attraction[1].xpath('./text()')[0]   # 景区所在地

            if location == name[:len(location)]:
                name = name[len(location):]

            level = '4A' if guizhou_idx == 0 else '5A'      # 级别
            level_time = attraction[7].xpath('./text()')[0][:4]  # 评级时间

            infos['贵州'].append([name, location, level+level_time+'年'])
    # ---------- 贵州 ----------

    # ---------- 北京 ----------
    beijing_url1 = "https://www.maigoo.com/public/mod/php/getpage.php" \
                   "?templateid=21758&append=1&dataid=2239737&ismobile=0&page=1&numshow=250&num=20&startid="

    beijing_url2 = "&blockac=zhishi&blockitid=180977&action=getpage&from="

    for i in range(0, 220, 20):
        beijing_url = beijing_url1 + str(i) + beijing_url2
        response = requests.get(url=beijing_url, headers=headers)
        page_text = response.text
        page_text = etree.HTML(page_text)
        attractions = page_text.xpath('.//tr[@class="li font14"]')
        for attraction in attractions:
            attraction_name = attraction[1].xpath('./text()')
            if not attraction_name:
                attraction_name = attraction[1].xpath('.//a[@target="_blank"]/text()')[0]

            if isinstance(attraction_name, list) and len(attraction_name) > 0:
                attraction_name = attraction_name[0]
            # 2. 景区所在地
            location = attraction[2].xpath('./text()')[0]

            # 3. 景区级别和评级时间
            attraction_level = attraction[3].xpath('./text()')[0]. \
                replace('（', '(').replace('）', ')').replace(' ', '')

            infos['北京'].append([attraction_name, location, attraction_level])
    # ---------- 北京 ----------

    for idx, (k, v) in enumerate(infos.items()):
        if not v:
            print(idx+1, k)

    return infos


def proce_excel(infos):
    for idx, (province_name, per_infos) in enumerate(infos.items()):
        wb = Workbook()
        ws = wb.active

        # 添加表头
        ws.append(['名称', '所在地', '级别', '评级时间'])

        for value in per_infos:
            # value : [景区名称, 景区所在地, 景区级别和评级时间]
            value[2] = value[2].replace('(', '').replace(')', '')
            if '年' == value[2][-1]:
                level_time = str(value[2][2:6])
                level = value[2][:2]
            elif '年' != value[2][-1] and '年' in value[2]:
                level_time = str(value[2].split('年')[0])
                level = value[2].split('年')[1]
            else:
                level_time = None
                level = value[2]
            if '等' in value[1]:
                value[1] = value[1][:-1]
            ws.append([value[0], value[1], level, level_time])
        save_path = save_root / (str(idx+1) + '_' + province_name + '.xlsx')
        wb.save(save_path)


def run():
    infos = main()
    proce_excel(infos=infos)


def guizhou():
    guizhou_url_4a = "https://www.guizhou.gov.cn/dcgz/hkgz/rmjd/202109/t20210914_70370942.html"
    guizhou_url_5a = "https://www.guizhou.gov.cn/dcgz/hkgz/rmjd/202109/t20210914_70370943.html"
    urls = [guizhou_url_4a, guizhou_url_5a]
    for guizhou_idx, i in enumerate(urls):
        response = requests.get(url=i, headers=headers)
        # 写成这种格式，防止中文乱码
        page_text = etree.HTML(response.content, parser=etree.HTMLParser(encoding='utf8'))
        # 去掉第一个
        attractions = page_text.xpath('.//div[@class="ue_table"]//tr')[1:]
        for attraction in attractions:
            name = attraction[2].xpath('./text()')[0]       # 景区名称
            location = attraction[1].xpath('./text()')[0]   # 景区所在地
            level = '4A' if guizhou_idx == 0 else '5A'      # 级别
            level_time = attraction[7].xpath('./text()')[0][:4]  # 评级时间

            print(name, location, level+level_time+'年')


def test_beijing():
    beijing_url1 = "https://www.maigoo.com/public/mod/php/getpage.php" \
                   "?templateid=21758&append=1&dataid=2239737&ismobile=0&page=1&numshow=250&num=20&startid="

    beijing_url2 = "&blockac=zhishi&blockitid=180977&action=getpage&from="

    for i in range(0, 220, 20):
        beijing_url = beijing_url1 + str(i) + beijing_url2
        response = requests.get(url=beijing_url, headers=headers)
        page_text = response.text
        page_text = etree.HTML(page_text)
        attractions = page_text.xpath('.//tr[@class="li font14"]')
        for attraction in attractions:
            attraction_name = attraction[1].xpath('./text()')
            if not attraction_name:
                attraction_name = attraction[1].xpath('.//a[@target="_blank"]/text()')[0]

            if isinstance(attraction_name, list) and len(attraction_name) > 0:
                attraction_name = attraction_name[0]
            # 2. 景区所在地
            location = attraction[2].xpath('./text()')[0]

            # 3. 景区级别和评级时间
            attraction_level = attraction[3].xpath('./text()')[0]. \
                replace('（', '(').replace('）', ')').replace(' ', '')

            print(attraction_name, location, attraction_level)


def parse_excel():
    nums = 0
    none_nums = 0
    for file_path in save_root.iterdir():
        workbook = load_workbook(filename=file_path)
        sheet = workbook["Sheet"]
        rows = sheet.max_row
        nums += (rows-1)
        for i in range(1, rows):
            pos = 'D' + str(i)
            cell_value = sheet[pos].value
            if cell_value is None:
                none_nums += 1

    print(f"nums = {nums} none_nums = {none_nums}")


def main2():
    url = "https://www.maigoo.com/goomai/197792.html"
    response = requests.get(url=url, headers=headers)
    page_text = response.text

    # xpath解析网页源码
    page_text = etree.HTML(page_text)

    provinces = page_text.xpath(
        '//div[@class="md_citiaothemetop10 md_citiaothemetop10_clos"]/div')[0].xpath('.//li')

    print(len(provinces))


if __name__ == "__main__":
    # run()
    main2()

